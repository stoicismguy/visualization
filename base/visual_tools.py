import openpyxl
from django.http import HttpResponse
from openpyxl.utils import get_column_letter

class Node:
    def __init__(self, value, position, max_depth=None, isLastinRow=False, print_exclude=False):
        self.value = value
        self.children = dict()
        self.position = position
        self.max_depth = max_depth
        self.isLastInRow = isLastinRow
        self.print_exclude = print_exclude #все норм, для того, чтоб mainNode не выводилась

    def __str__(self):
        if self.value is None:
            return ""
        return self.value

    def add_child(self, note: 'Node') -> None:
        if note.value in self.children.keys():
            d1 = self.children[note.value].children
            d2 = note.children
            common_keys = d1.keys() & d2.keys()
            if len(common_keys) == 0:
                self.children[note.value].children = {**self.children[note.value].children, **note.children}
            else:
                for key in common_keys:
                    for child in d2[key].children.values():
                        d1[key].add_child(child)
        else:
            self.children[note.value] = note

    def get_children(self):
        return self.children.values()
    

    def is_child_empty(self):
        if self.isLastInRow:
            return True
        if len(self.get_children()) != 1 or (len(self.get_children()) == 1 and list(self.get_children())[0].value is not None):
            return False     
        return list(self.get_children())[0].is_child_empty()
        
    
    def self_html_representation(self):
        # if self.value is None and self.isLastInRow:
        #     return ""
        # elif self.value is None:
        #     html = f"<li><p style='visibility:hidden;'>{self}</p>"
        # else:
        #     html = f"<li><p>{self}</p>"


        # if self.value is None and not self.is_child_empty():
        #     html = f"<li><p style='visibility:hidden;'>{self}</p>"
        # elif self.value is None and self.isLastInRow:
        #     return ""
        # else:
        #     html = f"<li><p>{self}</p>"


        if self.value is None and self.is_child_empty():
            html = f"<li>"
        elif self.value is None:
            html = f"<li><p style='visibility:hidden;'>{self}</p>"
        else:
            html = f"<li><p>{self}</p>"

        # if self.isLastInRow: html += "(last)"

        if self.print_exclude:
            html = "<li>"
        

        children = self.get_children()
        if len(children) != 0:
            html += f"<ul class='{self.position+1}'>"
        for child in children:
            html += child.self_html_representation()
        if len(children) != 0:
            html += "</ul>"
        html += "</li>"
        return html
    

    def print(self, level=0):
        print("-" * level, self.value)
        for i in self.children.values():
            i.print(level+2)


def get_active_file(file_link):
    wb = openpyxl.load_workbook(file_link)
    file = wb.active
    return file


def generate_visualization_tree(file) -> Node:
    mainNode = Node("", 0, 0, print_exclude=True)
    for line in file.rows:
        isFirst = True
        prevNote = None
        startNode = None
        node_position = 1
        for v in line:
            currentNote = Node(v.value, node_position)      
            node_position += 1
            mainNode.max_depth = max(mainNode.max_depth, node_position)
            if isFirst:
                startNode = currentNote
                prevNote = currentNote
                isFirst = False
            else:
                prevNote.add_child(currentNote)
                prevNote = currentNote   
        prevNote.isLastInRow = True
        mainNode.add_child(startNode)
    return mainNode


def delete_rows_filtering(active_table, filters):
    for i in range(active_table.max_row, 0, -1):
        row = active_table[i]
        for index, values in filters.items():
            # print(index, values, row)
            if str(row[index].value) not in values:
                # active_table.delete_rows(i)
                # print(index+1, active_table.max_column)
                for new_i in range(index+1, active_table.max_column+1):
                    active_table[f'{get_column_letter(new_i)}{i}'].value = None
                break


def delete_columns_by_number(active_table, columns):
    for c in reversed(sorted(list(columns))):
        active_table.delete_cols(int(c))








