from django.shortcuts import render, redirect
from django.http import HttpResponse
import json
from rest_framework.views import APIView
from .forms import DocumentForm
from .models import Document
from .visual_tools import generate_visualization_tree, get_active_file, delete_rows_filtering, delete_columns_by_number
from openpyxl.utils import get_column_letter

# Create your views here.
def index(request):
    return render(request, 'index.html')



class Index(APIView):
    def get(self, request):
        form = DocumentForm()
        return render(request, 'index.html', {"form": form, "user":request.user})

    def post(self, request):
        print("FORM")
        form = DocumentForm(request.POST, request.FILES)
        print(request.FILES)
        if form.is_valid():
            newdoc = Document(user=request.user,
                              docfile=request.FILES.get('docfile'))
            newdoc.save()
        form = DocumentForm()
        return redirect("index")
        # return render(request, 'index.html', {"form": form})
    

def files_view(request):
    documents = Document.objects.filter(user=request.user).order_by('-time')
    return render(request, "my_files.html", {"documents": documents})



class OutColumn:
    def __init__(self, values):
        self.values = values
        


def visualization_view(request, file_id):
    file = Document.objects.get(id=file_id)
    open_file = get_active_file(file.docfile.path)
    post = request.GET.keys()

    on_checkboxes = [x for x in post if x.isdigit()]

    on_checkboxes_out = list(map(int, list(on_checkboxes)))
    all_checkboxes_out = []
    for i in range(open_file.max_column):
        values = []
        for j in open_file[get_column_letter(i+1)]:
            if j.value is not None:
                values.append(j.value)
        values = sorted(set(values))
        c = OutColumn(values)
        all_checkboxes_out.append(c)

    filters = {}

    for i in on_checkboxes:
        all_values = dict(request.GET).get(f"{i}v")
        filters[int(i)-1] = set(all_values)
        
    delete_rows_filtering(open_file, filters)
    on_checkboxes=set([x-1 for x in on_checkboxes_out])

    if len(set(on_checkboxes_out)) != 0:
        set_to_delete = reversed(list(set(list(range(open_file.max_column))) - set(on_checkboxes)))
        for i in set_to_delete:
            open_file.delete_cols(i+1, 1)
    else:
        on_checkboxes_out = [x+1 for x in range(open_file.max_column)]

    tree_node = generate_visualization_tree(open_file)
    return render(request, "visual.html", {"node": tree_node,
                                           "file_name": file.docfile.name,
                                           "user": request.user,
                                           "columns":all_checkboxes_out,
                                           "checked": on_checkboxes_out,
                                           "file_id_t": file_id})



def get_node_html_by_filters(request, file_id):
    file = Document.objects.get(id=file_id)
    open_file = get_active_file(file.docfile.path)

    data = request.POST
    data = dict(request.POST)
    # print(data)

    non_active_columns = json.loads(data.get("unactive_buttons")[0])
    # print(non_active_columns)

    filters = {}
    for i in range(len(data)-2):
        d = data.get(f"{i}[]")
        if d is None:
            filters[i] = set([])
        else:
            filters[i] = set(d)

    delete_rows_filtering(open_file, filters)
    delete_columns_by_number(open_file, non_active_columns)
    node = generate_visualization_tree(open_file)
    a = node.self_html_representation()
    return HttpResponse(a)



def delete_file(request, file_id):
    doc = Document.objects.get(id=file_id, user=request.user)
    if doc is not None:
        doc.docfile.delete()
        doc.delete()
        # doc.save()
    return redirect("files")